from openpyxl import Workbook, load_workbook

wb = Workbook()

ws_test = wb.create_sheet(title='TestSheet')

count = 0
sum = 0

# for row in ws_test["B2:E5"]: <- わかりやすい書き方
for row in ws_test.iter_rows(min_row=2, min_col=2, max_row=5, max_col=5):
    for cell in row:
        count += 1
        sum += count
        cell.value = count

ws_test.cell(8, 2, "平均")
ws_test.cell(8, 3, "合計")
ws_test.cell(9, 2, sum / count)
ws_test.cell(9, 3, sum)

ws_input = wb.create_sheet(title="DataInput")

for i in range(1, 11):
    ws_input.cell(i, 1, i)
    ws_test.cell(i + 1, 6, i + 10)

wb.save("./python_module/ExcelResult_文台湖.xlsx")