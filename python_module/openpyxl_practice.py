from openpyxl import Workbook, load_workbook

# wb = Workbook()

wb = Workbook()
ws = wb.create_sheet(title='Sheet1')

for row in ws.rows:
    for cell in row:
        print(cell.value)

# ワークシートからデータ取得
""" print(ws["A1"].value)
print(ws["B2"].value)
print(ws.cell(1, 1).value) 
print(ws.cell(2, 2).value)
print(ws.cell(3, 2).value) # データがない場合はNone """


#ワークシートの編集
""" # ws["A1"] = 1
ws.cell(1, 1, 1)

ws["B1"] = 2
ws.cell(1, 2, 2)

ws["A2"] = 3
ws.cell(2, 1, 3)

ws["B2"] = 4
ws.cell(2, 2, 4)

row_data = [1, 2, 3, 4]
ws.append(row_data)
ws.append([1])
ws.append([1, 2, 3])
 """

for row in ws.iter_rows(min_row=1, min_col=1, max_row=2, max_col=2):
    for cell in row:
        print(cell.value)

wb.save("./python_module/test.xlsx")

