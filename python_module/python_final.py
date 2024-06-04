import pyodbc
import os
import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

load_dotenv()


SERVER = os.getenv("SERVER")
UID = os.getenv("DB_USER")
PWD = os.getenv("DB_PASSWORD")
DATABASE = os.getenv("DATABASE")

connect_info = 'Driver={}; Server={}; uid={}; pwd={}; Database={};'.format("SQL Server", SERVER, UID, PWD, DATABASE)

# データベース接続（①）
conn = pyodbc.connect(connect_info)

cursor = conn.cursor()

# テーブル作成（②）
# cursor.execute("CREATE TABLE movies (id int not null identity(1, 1) primary key, Release_Date datetime, Title varchar(2000), Overview varchar(2000), Popularity float, Vote_Count int, Vote_Average float, Original_Language varchar(2000), Genre varchar(2000), Poster_Url varchar(2000))")

# CSVを読み込む
# df = pd.read_csv("./python_module/mymoviedb.csv")

# NULL処理
# df = df.dropna(how="any")

# Dataframeをデータベースに入れる
# for index, row in df.iterrows():
#     columns_name = "Release_Date, Title, Overview, Popularity, Vote_Count, Vote_Average, Original_Language, Genre, Poster_Url"
#     release_date = row["Release_Date"]
#     title = row["Title"]
#     overview = row["Overview"]
#     popularity = row["Popularity"]
#     vote_count = row["Vote_Count"]
#     vote_average = row["Vote_Average"]
#     original_language = row["Original_Language"]
#     genre = row["Genre"]
#     poster_url = row["Poster_Url"]
#     cursor.execute("INSERT INTO movies ({columns_name}) VALUES ({release_date}, {title}, {overview}, {popularity}, {vote_count}, {vote_average}, {original_language}, {genre}, {poster_url})".format(columns_name, release_date, title, overview, popularity, vote_count, vote_average, original_language, genre, poster_url))

# データベースからすべてのデータ取得
df_all = pd.read_sql("SELECT * FROM movies", conn)

# データベースから言語ごとの映画の数を取得
df_lan = df_all.groupby('Original_Language').count()
# df_lan = pd.read_sql("SELECT Original_Language, count(*) as Count FROM movies GROUP BY Original_Language", conn)
print(df_lan)

# データベースからジャンルごとの評価平均値を取得
df_genre = df_all.groupby('Genre').mean("Vote_Average").Vote_Average
print(df_genre)
# df_genre = pd.read_sql("SELECT Genre, avg(Vote_Average) as 'Average_Score' FROM movies GROUP BY Genre", conn)

# for row in cursor:
#     print(row)

# conn.commit()

conn.close()

wb = Workbook()

# Language Countsシート作成
ws_lan = wb.create_sheet("Language Counts")

ws_lan.cell(1, 1, "ID")
ws_lan.cell(1, 2, "Original_Language")
ws_lan.cell(1, 3, "Count")

# Language Countsシートにデータを入れる
for i in range(2, len(df_lan)):
        ws_lan.cell(i, 1, i - 1)
        ws_lan.cell(i, 2, df_lan["Original_Language"][i])
        ws_lan.cell(i, 3, df_lan["Count"][i])

# Genre Ratingsシート作成
ws_genre = wb.create_sheet("Genre Ratings")

ws_genre.cell(1, 1, "ID")
ws_genre.cell(1, 2, "Genre")
ws_genre.cell(1, 3, "Average Score")

# Genre Ratingsシートにデータを入れる
for i in range(2, len(df_genre)):
        ws_genre.cell(i, 1, i - 1)
        ws_genre.cell(i, 2, df_genre["Genre"][i])
        ws_genre.cell(i, 3, df_genre["Average_Score"][i])

wb.save("./python_module/AnalyzeMovies_文台湖.xlsx")